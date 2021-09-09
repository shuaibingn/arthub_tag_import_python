from openpyxl import load_workbook

from utils.const import is_sync


class XLSX(object):

    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(filename=filename)
        self.ws = self.wb.get_sheet_by_name(self.wb.sheetnames[0])

    def read_data(self) -> list:
        res = []
        i = 0
        for row in self.ws.rows:
            i += 1
            j = 0
            d = {}
            for cell in row:
                j += 1
                k = self.ws.cell(row=1, column=j).value
                if i != 1 and k == is_sync:
                    d["index"] = [i, j]

                if i != 1:
                    v = cell.value
                    if isinstance(v, float):
                        v = int(v)
                    d[k] = str(v)

            if d and d.get(is_sync) == "否":
                res.append(d)

        return res

    def write_data(self, data: list):
        for d in data:
            index = d.get("index")
            self.ws.cell(index[0], index[1], "是")
        self.wb.save(filename=self.filename)
